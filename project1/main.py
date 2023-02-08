import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}


for products_row in range(2, product_list.max_row + 1):

    suppiler_name = product_list.cell(products_row, 4).value
    inventory = product_list.cell(products_row, 2).value
    price = product_list.cell(products_row, 3).value
    product_number = product_list.cell(products_row, 1).value
    inventory_price = product_list.cell(products_row, 5)

    if suppiler_name in products_per_supplier:
        products_per_supplier[suppiler_name] += 1
    else:
        products_per_supplier[suppiler_name] = 1

    if suppiler_name in total_value_per_supplier:
        total_value_per_supplier[suppiler_name] += inventory * price
    else:
        total_value_per_supplier[suppiler_name] = inventory * price

    if inventory < 10:
        products_under_10_inv[product_number] = inventory

    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)

inv_file.save("inventory_with_total_value.xlsx")
